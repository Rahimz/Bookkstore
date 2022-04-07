from django.shortcuts import render, get_object_or_404
from django.conf import settings
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from datetime import datetime
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
from django.core.cache import cache

from openpyxl import load_workbook
from io import BytesIO

from .models import File as FileObject, ImportSession
from django.core.files import File

from products.models import Product, Category, Publisher
from tools.fa_to_en_num import number_converter


def list_of_files(request):
    files_list = list(FileObject.objects.all().values_list('pk', flat=True))[-1:]
    files = FileObject.objects.filter(id__in=files_list)
    return render(
        request,
        'files/list.html',
        {
        'files': files,
        }
    )


@staff_member_required
def add_to_database(request, file_slug=None):
    """
    This Function is designed to import datat from excel.
    The imported datat is very different, so we grab every files
    and check the headers and import file in a manual method.
    """
    files = FileObject.objects.all()

    file_object = None
    row = None
    temp_object = None

    # use for reports
    number_of_added_object = 0

    # erroe report
    error_report = {}

    # grab a list of available barcode numbers
    barcode_number_list =Product.objects.all().values_list('barcode_number', flat=True)
    if file_slug:
        file_object = get_object_or_404(FileObject, slug=file_slug)

        myfile = File(file_object)

        path = file_object.file.path
        with open(path, 'rb') as f:
            # Load excel workbook
            wb = load_workbook(f)
            ws = wb.active
            row_count = ws.max_row

            # a loop for scrape the excel file
            category = Category.objects.get(name='other')

            # we make a label for each import session to  control it
            current_import_session = ImportSession.objects.create(user=request.user,)

            # read the data in cells
            for i in range(2, row_count):

                row = ws['A'+str(i):'M'+str(i)]

                temp_object = Product(
                    name = ws['M' + str(i)].value,
                    stock = ws['L' + str(i)].value,
                    author = ws['K' + str(i)].value,
                    translator = ws['J' + str(i)].value,
                    price = ws['I' + str(i)].value,
                    publish_year = ws['H' + str(i)].value,
                    edition = ws['G' + str(i)].value,
                    publisher = ws['F' + str(i)].value,
                    isbn = ws['E' + str(i)].value,
                    product_type = ws['D' + str(i)].value,
                    cover_type = ws['C' + str(i)].value,
                    barcode_number = ws['B' + str(i)].value,
                    size = ws['A' + str(i)].value,
                    category = category,
                    import_session=current_import_session,
                )

                if temp_object.barcode_number not in barcode_number_list:
                    try:
                        temp_object.save()
                        number_of_added_object += 1
                    except:
                        error_report[ws['M' + str(i)]] = ws['M' + str(i)].value
                        messages.error(request, 'Error updating your user details')



        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    return render(request,
                  'files/upload_database.html',
                  {'files': files,
                   'number_of_added_object': number_of_added_object,
                   'error_report': error_report,
                   })


def add_new_book_to_database(request, file_slug, check='check'):
    """
    to add new-book of the bookstore to datebase
    """

    duplicate_isbn = []
    isbn_error = []
    update_product = 0
    number_of_added_object = 0

    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path

    isbn_list = Product.objects.filter(isbn__isnull=False).values_list('isbn', flat=True)
    name_list = Product.objects.values_list('name', flat=True)
    barcode_number_list = Product.objects.filter(barcode_number__isnull=False).values_list('barcode_number', flat=True)

    if check == 'add':

        # we make a label for each import session to  control it
        current_import_session = ImportSession.objects.create(user=request.user,)

    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):

            row = ws['A'+str(i):'M'+str(i)]
            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value if not None else (0,),
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,

            if page_number[0] == '':
                page_number = [0,]
            if price[0] == '----':
                price = [0,]


            if check == 'add':
                product = Product.objects.create(
                    name = name[0],
                    isbn = str(isbn[0]),
                    price = price[0],
                    page_number =page_number[0],
                    publisher = publisher[0],
                    stock = stock[0],
                    state = 'new',
                    available = True,
                    available_in_store = True,
                    available_online = True,
                    import_session=current_import_session,
                )
                number_of_added_object += 1
                update_product += 1

            # print(price, type(price))
            # print(page_number, type(page_number))
            if str(isbn[0]) in isbn_list:
                duplicate_isbn.append((i, str(isbn[0])))

            if str(name[0]) in name_list:
                duplicate_isbn.append((i, str(isbn[0])))

            if len(str(isbn[0]))!= 13:
                isbn_error.append((i, str(isbn[0])))


    current_import_session.quantity = number_of_added_object
    current_import_session.save()

    # print(len(duplicate_isbn))
    # print(len(set(duplicate_isbn)))
    return render(
        request,
        'files/check_update.html',
        {
            'file': myfile,
            # 'number_of_added_object': number_of_added_object,
            # 'error_report': error_report,
            'isbn_list': isbn_list,
            'isbn_error': isbn_error,
            'duplicate_isbn': duplicate_isbn,
            'row_count': row_count,
            'update_product': update_product,
         }
    )


def add_new_book_to_database_2(request, file_slug, check='check'):
    """
    to add new-book of the bookstore to datebase
    """

    duplicate_isbn = []
    isbn_error = []
    update_product = 0
    number_of_added_object = 0
    check_sum = 0
    price_change_list_ids = []
    product_updated_with_excel_ids = []
    new_book_to_update = []

    main_list_143 = []
    excel_isbn_in_db = []
    isbn_same_name_in_db = []
    different_name_with_same_isbn = []
    no_isbn_with_same_name_in_db = []
    excel_dic = []

    updated_product_to_keep_unchanged = Product.objects.filter(import_session__id='df0dca0d-4ba1-4b6f-b07a-8df5b8c3d121').values_list('isbn', flat=True)

    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path

    isbn_list = Product.objects.filter(available=True).exclude(import_session__id='df0dca0d-4ba1-4b6f-b07a-8df5b8c3d121').values_list('isbn', flat=True)
    name_list = Product.objects.filter(available=True).exclude(import_session__id='df0dca0d-4ba1-4b6f-b07a-8df5b8c3d121').values_list('name', flat=True)
    barcode_number_list = Product.objects.filter(barcode_number__isnull=False).values_list('barcode_number', flat=True)

    if check == 'add':

        # we make a label for each import session to  control it
        current_import_session = ImportSession.objects.create(user=request.user,)

    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):

            row = ws['A'+str(i):'M'+str(i)]
            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value if not None else (0,),
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,

            if page_number[0] == None:
                page_number = [0,]
            if price[0] == '----':
                price = [0,]


            # print(isbn[0])
            # print(len(isbn_list))
            if str(isbn[0]) not in updated_product_to_keep_unchanged:
                main_list_143.append(str(isbn[0]))
                # if str(isbn[0]) != 'None':
                database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
                # print('isbn in database', isbn[0], name[0])


                excel_isbn_in_db.append(str(isbn[0]))
                if name[0] == database_product.name:
                    if check == 'add':
                        database_product.stock += stock[0]
                        database_product.save()
                    # print('same isbn same name', isbn[0], name[0])
                    isbn_same_name_in_db.append(str(isbn[0]))
                else:
                    # print('different_name_with_same_isbn', isbn[0], name[0])
                    different_name_with_same_isbn.append(str(isbn[0]))
                    excel_dic.append((name[0], str(isbn[0]), stock[0]))

                    # new_book_to_update.append((isbn[0], name[0]))



            # # print(price, type(price))
            # # print(page_number, type(page_number))
            # if str(isbn[0])!='None' and (str(isbn[0]) in isbn_list):
            #     database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
            #
            #     if name[0] == database_product.name:
            #         product_updated_with_excel_ids.append(database_product.id)
            #         if price[0] != database_product.price:
            #             price_change_list_ids.append(database_product.id)
            #
            #
            #     if ws['B' + str(i)].value != database_product.name:
            #         duplicate_isbn.append((i, str(isbn[0]),ws['B' + str(i)].value, [Product.objects.filter(available=True).filter(isbn=str(isbn[0]))]))
            # else:
            #     if check == 'add':
            #         # print(name[0])
            #         product = Product.objects.create(
            #             name = name[0],
            #             isbn = str(isbn[0]),
            #             price = price[0],
            #             page_number =page_number[0],
            #             publisher = publisher[0],
            #             stock = stock[0],
            #             state = 'new',
            #             available = True,
            #             available_in_store = True,
            #             available_online = True,
            #             publish_year=None,
            #             import_session=current_import_session,
            #         )
            #         number_of_added_object += 1
            #         update_product += 1
            #

            if str(name[0]) in name_list:
                duplicate_isbn.append((i, str(isbn[0])))

            if len(str(isbn[0]))!= 13:
                isbn_error.append((i, str(isbn[0])))

    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()

    price_change_list = Product.objects.filter(pk__in=price_change_list_ids)
    product_updated_with_excel = Product.objects.filter(pk__in=product_updated_with_excel_ids)

    diff_name = Product.objects.filter(available=True).filter(isbn__in=different_name_with_same_isbn).order_by('isbn')
    # different_name_with_same_isbn = different_name_with_same_isbn.sort(key=lambda x:x[1])
    excel_dic = sorted(excel_dic, key=lambda x:x[1])

    print('should be as same as import session Q.',check_sum)
    print(len(updated_product_to_keep_unchanged) )
    # print(len(duplicate_isbn))
    # print(len(price_change_list))
    # print(len(product_updated_with_excel))
    # print(len(set(duplicate_isbn)))
    print('None' in updated_product_to_keep_unchanged)
    print('main_list_143: ', len(main_list_143))
    print('excel_dic: ', len(excel_dic))
    # print('excel_isbn_in_db: ', len(excel_isbn_in_db))
    # print('isbn_same_name_in_db: ', len(isbn_same_name_in_db))
    # print('different_name_with_same_isbn: ', len(different_name_with_same_isbn))
    print('no_isbn_with_same_name_in_db: ', len(no_isbn_with_same_name_in_db))
    return render(
        request,
        'files/check_update.html',
        {
            'file': myfile,
            # 'number_of_added_object': number_of_added_object,
            # 'error_report': error_report,
            'isbn_list': isbn_list,
            'isbn_error': isbn_error,
            'duplicate_isbn': duplicate_isbn,
            'row_count': row_count,
            'update_product': update_product,
            'price_change_list': price_change_list,
            'product_updated_with_excel': product_updated_with_excel,

            'excel_isbn_in_db': excel_isbn_in_db,
            'isbn_same_name_in_db': isbn_same_name_in_db,
            'different_name_with_same_isbn': different_name_with_same_isbn,
            'no_isbn_with_same_name_in_db': no_isbn_with_same_name_in_db,
            'main_list_143': main_list_143,
            'diff_name': diff_name,
            'excel_dic': excel_dic,
         }
    )


def update_isbn_duplicate(request, product_isbn):
    pass


def name_correction(request, file_slug, check='check'):
    # variables
    updated_product_count = 0
    more_than_one = 0
    not_updated = []
    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):

            row = ws['A'+str(i):'M'+str(i)]
            p_ex_id = ws['C' + str(i)].value,
            new_name = ws['F' + str(i)].value,


            try:
                database_product = Product.objects.get(pk=p_ex_id[0])
                if check == 'add':
                    updated_product_count +=1
                    database_product.name = new_name[0]
                    database_product.save()
            except:
                # print(p_ex_id[0], new_name[0])
                # print(new_name[0], type(new_name[0]))
                more_than_one +=1


    # print('updated_product_count', updated_product_count)
    # print('more_than_one', more_than_one)
    return render(
        request,
        'files/name_correction.html',
        {
            'file': myfile,
            'row_count': row_count,
            'updated_product_count': updated_product_count,
            'more_than_one': more_than_one,
         }
    )


def correction_140(request, file_slug, check='check'):
    data_check = []
    new_price_row = 0
    just_stock_update = 0
    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):

            row = ws['A'+str(i):'M'+str(i)]
            isbn = ws['C' + str(i)].value,
            price_1 = ws['D' + str(i)].value,
            stock_1 = ws['G' + str(i)].value,
            price = ws['H' + str(i)].value,
            price_change = ws['I' + str(i)].value,

            database_product = Product.objects.filter(available=True).get(isbn=isbn[0])

            # data_check.append((isbn[0], price_1[0], stock_1[0], price[0], price_change[0], database_product.pk, database_product.has_other_prices))
            if price_change[0]:
                # with two prices
                data_check.append(('*', isbn[0], price_1[0], stock_1[0], price[0], price_change[0], database_product.pk, database_product.has_other_prices))
            else:
                # just update stock
                data_check.append(( isbn[0], price_1[0], stock_1[0], price[0], price_change[0], database_product.pk, database_product.has_other_prices))


            if check == 'add':
                if price_change[0]:
                    # with two prices
                    database_product.has_other_prices = True
                    database_product.price_1 = price_1[0]
                    database_product.stock_1 = stock_1[0]
                    database_product.save()
                    new_price_row += 1
                else:
                    # just update stock
                    database_product.stock += stock_1[0]
                    database_product.save()
                    just_stock_update +=1

    return render(
        request,
        'files/correction_140.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'data_check': data_check,
            'new_price_row': new_price_row,
            'just_stock_update': just_stock_update,
         }
    )


def add_new_book_to_database_3(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []
    number_of_added_object = 0


    current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):

            row = ws['A'+str(i):'M'+str(i)]

            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,

            if not page_number[0]:
                page_number = [0,]

            try:
                database_product = Product.objects.filter(available=True).get(isbn=isbn[0])
                isbn_in_database.append(( isbn[0], price[0], stock[0], price[0],  database_product.pk, database_product.has_other_prices))
                if check == 'add':
                    database_product.stock += stock[0]
                    database_product.save()
            except:
                not_in_database.append(( isbn[0], price[0], stock[0],  ))
                if check == 'add':
                    product = Product.objects.create(
                        name = name[0],
                        isbn = str(isbn[0]),
                        price = price[0],
                        page_number =page_number[0],
                        publisher = publisher[0],
                        stock = stock[0],
                        state = 'new',
                        available = True,
                        available_in_store = True,
                        available_online = True,
                        import_session=current_import_session,
                    )
                    number_of_added_object += 1



    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    if current_import_session.quantity == 0:
        current_import_session.delete()
    return render(
        request,
        'files/add_series_3.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'isbn_in_database': isbn_in_database,
            'not_in_database': not_in_database,
            # 'new_price_row': new_price_row,
            # 'just_stock_update': just_stock_update,
         }
    )


def add_used_book_with_isbn(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []
    number_of_added_object = 0


    current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 15):
            in_db = False
            not_in_db = False

            row = ws['A'+str(i):'M'+str(i)]

            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            if not price[0]:
                price = [0,]

            # print(price[0], type(price[0]))

            if len(str(isbn[0])) == 13:
                isbn_9 = str(isbn[0])[3:-1]
            elif len(str(isbn[0])) == 10:
                isbn_9 = str(isbn[0])[:-1]
            elif len(str(isbn[0])) == 9:
                isbn_9 = str(isbn[0])
            else:
                isbn_9 = None

            try:
                # Check with 9 Char ISBN
                if isbn_9:
                    database_product = Product.objects.filter(available=True).get(isbn_9=isbn_9)
                else:
                    # Check with 13 char ISBN
                    database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
                isbn_in_database.append(( isbn[0], price[0], stock[0],  database_product.pk, database_product.has_other_prices))
                in_db = True
                # print(database_product.pk)
            except:
                not_in_database.append(( isbn[0], price[0], stock[0],  ))
                not_in_db = True
                # print(isbn[0], price[0], type(price[0]))


            if check == 'add':
                if in_db:
                    database_product.has_other_prices = True
                    database_product.price_used = price[0]
                    database_product.stock_used = stock[0]
                    database_product.save()
                if not_in_db:
                    product = Product.objects.create(
                        name = name[0],
                        isbn = str(isbn[0]),
                        page_number = page_number[0],
                        publisher = publisher[0],
                        price = 0,
                        stock = 0,
                        has_other_prices = True,
                        price_used = price[0],
                        stock_used = stock[0],
                        state = 'new',
                        available = True,
                        available_in_store = True,
                        available_online = True,
                        import_session=current_import_session,
                    )
                    number_of_added_object += 1



    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()
    print('isbn_in_database', len(isbn_in_database))
    print('not_in_database', len(not_in_database))
    return render(
        request,
        'files/add_used_isbn.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'isbn_in_database': isbn_in_database,
            'not_in_database': not_in_database,

         }
    )


def add_used_book_with_no_isbn(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []
    number_of_added_object = 0

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        # for i in range(2, 3):
        for i in range(2, row_count):
            row = ws['A'+str(i):'M'+str(i)]

            name = ws['B' + str(i)].value,
            # isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            if not price[0]:
                price = [0,]

            # print(price[0], type(price[0]))


            if check == 'add':
                product = Product.objects.create(
                    name = name[0],
                    # isbn = str(isbn[0]),
                    page_number = page_number[0],
                    publisher = publisher[0],
                    price = 0,
                    stock = 0,
                    has_other_prices = True,
                    price_used = price[0],
                    stock_used = stock[0],
                    state = 'new',
                    available = True,
                    available_in_store = True,
                    available_online = True,
                    import_session=current_import_session,
                )
                number_of_added_object += 1
            else:
                number_of_added_object += 1



    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()

    return render(
        request,
        'files/add_used_no_isbn.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'number_of_added_object': number_of_added_object,


         }
    )


def add_new_book_to_database_4(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []

    number_of_added_object = 0
    new_price = []
    new_row_price = []

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        # for i in range(2, 3):
        for i in range(2, row_count):
            row = ws['A'+str(i):'M'+str(i)]

            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            if not price[0]:
                price = [0,]

            if len(str(isbn[0])) == 13:
                isbn_9 = str(isbn[0])[3:-1]
            elif len(str(isbn[0])) == 10:
                isbn_9 = str(isbn[0])[:-1]
            elif len(str(isbn[0])) == 9:
                isbn_9 = str(isbn[0])

            try:
                database_product = Product.objects.filter(available=True).get(isbn_9=isbn_9)
                isbn_in_database.append(( isbn[0], price[0], stock[0], price[0],  database_product.pk, database_product.has_other_prices))
                if check == 'add':
                    # these are the same with databse and
                    # just should update stock
                    database_product.stock += stock[0]
                    database_product.save()

                if int(price[0]) != int(database_product.price):
                    # prices in excel and databse are not the same
                    if int(database_product.price) == 0 :

                        # we have this books as used and their main price is empty
                        new_row_price.append(( 'excel',isbn[0], name[0], price[0], publisher[0], 'product', int(database_product.price) ,'other prices', database_product.has_other_prices))
                        if check == 'add':
                            # to change 9 digits isbn to 13
                            if len(database_product.isbn) < len(str(isbn[0])):
                                database_product.isbn = str(isbn[0])

                            database_product.price = price[0]
                            database_product.stock = stock[0]
                            database_product.save()

                    else:
                        # these product should have a new price upper than what we have in database
                        new_price.append(( 'excel',isbn[0], name[0], price[0], publisher[0], 'product', int(database_product.price) ,'other prices', database_product.has_other_prices))
                        if check == 'add':
                            database_product.name = name[0]
                            database_product.stock_1 = database_product.stock
                            database_product.price_1 = database_product.price
                            database_product.stock = stock[0]
                            database_product.price = price[0]
                            database_product.has_other_prices = True
                            database_product.save()

            except:
                not_in_database.append(( isbn[0], price[0], stock[0],  ))
                if check == 'add':
                    product = Product.objects.create(
                        name = name[0],
                        isbn = str(isbn[0]),
                        price = price[0],
                        page_number =page_number[0],
                        publisher = publisher[0],
                        stock = stock[0],
                        state = 'new',
                        available = True,
                        available_in_store = True,
                        available_online = True,
                        import_session=current_import_session,
                    )
                    number_of_added_object += 1


    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()

    return render(
        request,
        'files/add_series_4.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'number_of_added_object': number_of_added_object,
            'isbn_in_database': isbn_in_database,
            'not_in_database': not_in_database,
            'new_price': new_price,
            'new_row_price': new_row_price,
         }
    )

def add_crafts(request, file_slug, check='check'):
    in_database = []
    number_of_added_object = 0

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        # for i in range(2, 3):
        for i in range(2, row_count):
            row = ws['A'+str(i):'M'+str(i)]

            barcode_number = ws['A' + str(i)].value,
            name = ws['B' + str(i)].value,
            stock = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            category = ws['E1'].value,

            # print(stock[0], type(stock[0]))
            # print(price[0], type(price[0]))
            in_database.append((category[0], name[0], price[0], stock[0]))

            if check == 'add':
                Product.objects.create(
                    barcode_number = barcode_number[0],
                    name = name[0],
                    stock = stock[0],
                    price = price[0],
                    craft_category = category[0],
                    product_type = 'craft',
                    publish_year = None,
                    available_online = False,
                )

                number_of_added_object += 1


    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()

    return render(
        request,
        'files/add_crafts.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'in_database': in_database,
         #    'number_of_added_object': number_of_added_object,
         #    'not_in_database': not_in_database,
         #    'new_price': new_price,
         #    'new_row_price': new_row_price,
         }
    )


def used_book_update(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []
    number_of_added_object = 0
    no_price_for_new_book = []

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 15):
            in_db = False
            not_in_db = False

            row = ws['A'+str(i):'M'+str(i)]

            name = ws['B' + str(i)].value,
            isbn = ws['C' + str(i)].value,
            price = ws['D' + str(i)].value,
            page_number = ws['E' + str(i)].value,
            publisher = ws['F' + str(i)].value,
            stock = ws['G' + str(i)].value,
            price_of_new_book = ws['H' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            if not price[0]:
                price = [0,]

            # print(price[0], type(price[0]))
            if price_of_new_book[0] == '-':
                no_price_for_new_book.append(( name[0],isbn[0], price[0], price_of_new_book[0] ))

            if len(str(isbn[0])) == 13:
                isbn_9 = str(isbn[0])[3:-1]
            elif len(str(isbn[0])) == 10:
                isbn_9 = str(isbn[0])[:-1]
            elif len(str(isbn[0])) == 9:
                isbn_9 = str(isbn[0])
            else:
                isbn_9 = None

            try:
                # Check with 9 Char ISBN
                if isbn_9:
                    database_product = Product.objects.filter(available=True).get(isbn_9=isbn_9)
                else:
                    # Check with 13 char ISBN
                    database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
                isbn_in_database.append(( isbn[0], price[0], stock[0],  database_product.pk, database_product.has_other_prices))
                in_db = True


                # print(database_product.pk)
            except:
                not_in_database.append(( isbn[0], price[0], stock[0],  ))
                not_in_db = True
                # print(isbn[0], price[0], type(price[0]))


            if check == 'add':
                if in_db:
                    if price_of_new_book[0] == '-':
                        # we add note to say ther is no new version of this book
                        database_product.about = '*'
                    else:
                        database_product.price = price_of_new_book[0]

                    if price[0]:
                        # we add price if there is any used price
                        database_product.price_used = price[0]

                    database_product.has_other_prices = True
                    database_product.save()

                if not_in_db:
                    product = Product.objects.create(
                        name = name[0],
                        isbn = str(isbn[0]),
                        page_number = page_number[0],
                        publisher = publisher[0],
                        price = price_of_new_book[0],
                        stock = 0,
                        has_other_prices = True,
                        price_used = price[0],
                        stock_used = stock[0],
                        state = 'new',
                        available = True,
                        available_in_store = True,
                        available_online = True,
                        import_session=current_import_session,
                    )
                    number_of_added_object += 1



    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()
    print('isbn_in_database', len(isbn_in_database))
    print('not_in_database', len(not_in_database))
    print('no_price_for_new_book', len(no_price_for_new_book))
    return render(
        request,
        'files/used_book_update.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'isbn_in_database': isbn_in_database,
            'not_in_database': not_in_database,
            'no_price_for_new_book': no_price_for_new_book,

         }
    )


def add_used_book_with_isbn_2(request, file_slug, check='check'):
    isbn_in_database = []
    not_in_database = []
    number_of_added_object = 0

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 15):
            in_db = False
            not_in_db = False

            row = ws['A'+str(i):'D'+str(i)]

            isbn = ws['A' + str(i)].value,
            name = ws['B' + str(i)].value,
            publisher = ws['C' + str(i)].value,
            page_number = ws['D' + str(i)].value,
            stock = ws['E' + str(i)].value,
            # price = ws['D' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            # if not price[0]:
            #     price = [0,]

            # print(isbn[0], type(isbn[0]))

            # if len(str(isbn[0])) == 13:
            #     isbn_9 = str(isbn[0])[3:-1]
            # elif len(str(isbn[0])) == 12:
            #     isbn_9 = str(isbn[0])[3:]
            # elif len(str(isbn[0])) == 10:
            #     isbn_9 = str(isbn[0])[:-1]
            # elif len(str(isbn[0])) == 9:
            #     isbn_9 = str(isbn[0])
            # else:
            #     isbn_9 = None

            try:
                # # Check with 9 Char ISBN
                # if isbn_9:
                #     database_product = Product.objects.filter(available=True).get(isbn_9=isbn_9)
                # else:
                #     # Check with 13 char ISBN
                database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
                isbn_in_database.append(( isbn[0],  stock[0],  database_product.pk, database_product.has_other_prices))
                in_db = True
                # print(database_product.pk)
            except:
                not_in_database.append(( isbn[0], stock[0],  name[0]))
                not_in_db = True
                # print(isbn[0], price[0], type(price[0]))


            if check == 'add':
                if in_db:
                    database_product.has_other_prices = True
                    # database_product.price_used = price[0]
                    database_product.stock_used += stock[0]
                    database_product.save()
                if not_in_db:
                    product = Product.objects.create(
                        name = name[0],
                        isbn = str(isbn[0]),
                        page_number = page_number[0],
                        publisher = publisher[0],
                        price = 0,
                        stock = 0,
                        has_other_prices = True,
                        price_used = 0,
                        stock_used = stock[0],
                        state = 'new',
                        product_type='book',
                        available = True,
                        available_in_store = True,
                        available_online = True,
                        import_session=current_import_session,
                    )
                    number_of_added_object += 1



    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    print('isbn_in_database', len(isbn_in_database))
    print('not_in_database', len(not_in_database))
    return render(
        request,
        'files/add_used_isbn.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'isbn_in_database': isbn_in_database,
            'not_in_database': not_in_database,

         }
    )


def add_used_book_with_no_isbn_2(request, file_slug, check='check'):
    name_in_database = []
    name_not_in_database = []
    number_of_added_object = 0
    number_of_duplicated_name = 0
    products = None

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        # for i in range(2, 3):
        for i in range(2, row_count):
            row = ws['A'+str(i):'D'+str(i)]

            name = ws['A' + str(i)].value,
            publisher = ws['B' + str(i)].value,
            page_number = ws['C' + str(i)].value,
            stock = ws['D' + str(i)].value,
            # isbn = ws['C' + str(i)].value,
            # price = ws['D' + str(i)].value,
            # print(price[0], type(price[0]))
            if not page_number[0]:
                page_number = [0,]
            price = [0,]

            # print(price[0], type(price[0]))
            # we want to check the name of product in databse before add them
            try:
                products = Product.objects.filter(name=name[0])
            except:
                pass
            if products:
                if len(products) == 1:
                    product = products.first()
                    name_in_database.append((name[0], product.isbn))
                else:
                    number_of_duplicated_name += 1
                    print('duplicated', product.name)
            else:
                name_not_in_database.append(name[0])

            if check == 'add':
                product = Product.objects.create(
                    name = name[0],
                    # isbn = str(isbn[0]),
                    page_number = page_number[0],
                    publisher = publisher[0],
                    price = 0,
                    stock = 0,
                    has_other_prices = True,
                    price_used = price[0],
                    stock_used = stock[0],
                    state = 'new',
                    product_type='book',
                    available = True,
                    available_in_store = True,
                    available_online = True,
                    import_session=current_import_session,
                )
                number_of_added_object += 1




    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()

    print('name_in_database', len(name_in_database),)
    print('name_not_in_database', len(name_not_in_database),)
    print('number_of_duplicated_name', number_of_duplicated_name)
    return render(
        request,
        'files/add_used_no_isbn.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'number_of_added_object': number_of_added_object,
            'name_in_database': name_in_database,
            'name_not_in_database': name_not_in_database,

         }
    )



def used_book_update_2(request, file_slug, check='check'):
    duplicate_price = []
    not_in_database = []
    number_of_added_object = 0
    remove_book = []

    if check == 'add':
        current_import_session = ImportSession.objects.create(user=request.user,)

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 15):
            in_db = False
            not_in_db = False

            row = ws['A'+str(i):'M'+str(i)]

            product_id = ws['B' + str(i)].value,
            name = ws['C' + str(i)].value,
            isbn = ws['D' + str(i)].value,
            price_used = ws['H' + str(i)].value,

            # print(price[0], type(price[0]))

            if not price_used[0]:
                price_used = [0,]

            # print(price[0], type(price[0]))
            if price_used[0] == 'xx':
                remove_book.append( (product_id[0], name[0],isbn[0]) )


            database_product = Product.objects.get(pk=product_id[0])
            # check for duplicate prices
            if database_product.price_used and database_product.price_used != price_used[0]:
                duplicate_price.append((product_id, database_product.price_used, price_used[0]))


            if check == 'add':
                if price_used[0] == 'xx':
                    database_product.available = False
                    database_product.available_in_store = False
                    database_product.available_online = False
                    database_product.save()
                else:
                    database_product.price_used = price_used[0]
                    database_product.has_other_prices = True
                    database_product.save()
                    number_of_added_object += 1

    if check == 'add':
        current_import_session.quantity = number_of_added_object
        current_import_session.save()
    # if current_import_session.quantity == 0:
    #     current_import_session.delete()
    # print('isbn_in_database', len(isbn_in_database))
    print('duplicate_price', len(duplicate_price))
    print('remove_book', len(remove_book))
    return render(
        request,
        'files/used_book_update_2.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'duplicate_price': duplicate_price,
            'not_in_database': not_in_database,
            'remove_book': remove_book,

         }
    )


def duplicate_book_1(request, file_slug, check='check'):
    has_main_stock = []
    # not_in_database = []
    # number_of_remove_object = 0
    remove_book = []

    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 500):

            # row = ws['A'+str(i):'M'+str(i)]

            product_id = ws['B' + str(i)].value,
            name = ws['C' + str(i)].value,
            remove = ws['D' + str(i)].value,
            main_stock = ws['J' + str(i)].value,
            flow = ws['N' + str(i)].value,

            if not remove[0]:
                remove = ['',]

            if flow[0] == None:
                flow = ['',]

            # the rows marked to eliminated
            if '/' in  remove[0]:
                database_product = Product.objects.get(pk=product_id[0])
                # we check the stock and flow of selected rows
                if not main_stock[0] and not database_product.stock and not database_product.stock_1 and not database_product.stock_2 and not flow[0]:
                    remove_book.append( (product_id[0], name[0]) )
                    if check == 'add':
                        database_product.available = False
                        database_product.available_in_store = False
                        database_product.available_online = False
                        if not database_product.admin_note:
                            database_product.admin_note = str(datetime.now().date()) + " remove because of duplication, "
                        else:
                            database_product.admin_note += " - " + str(datetime.now().date()) + " remove because of duplication, "
                        database_product.save()

                else:
                    has_main_stock.append(
                        (
                            product_id[0],
                            name[0],
                            flow[0],
                            main_stock,
                            database_product.stock,
                            database_product.stock_1,
                            database_product.stock_2,
                            flow[0]
                        )
                    )
                    if check == 'add':
                        if not database_product.admin_note:
                            database_product.admin_note = str(datetime.now().date()) + 'It should be removed but has flow or main stock'
                        else:
                            database_product.admin_note += " - " + str(datetime.now().date()) + 'It should be removed but has flow or main stock'
                        database_product.save()

            result_to_remove = [item[0] for item in has_main_stock]



    print('has_main_stock', len(has_main_stock))
    print('remove_book', len(remove_book))
    return render(
        request,
        'files/duplicate_book_1.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'has_main_stock': has_main_stock,
            'remove_book': remove_book,
            'result_to_remove': result_to_remove,

         }
    )


def check_category(request, file_slug, check='check', start=None, end=None):
    main_categories = []
    sub_categories = []
    not_found_product = []
    no_isbn = []
    duplicate_name = []


    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        # for i in range(2, row_count):
        for i in range(start, end):
            database_product = None
            database_products = None
            # row = ws['A'+str(i):'M'+str(i)]

            name = ws['A' + str(i)].value,
            isbn = ws['B' + str(i)].value,
            main_category = ws['C' + str(i)].value,
            sub_category = ws['D' + str(i)].value,
            publisher = ws['E' + str(i)].value,
            pages = ws['F' + str(i)].value,

            isbn = [number_converter(str(isbn[0]))]
            if 'x' in (isbn[0]):
                isbn = [number_converter(str(isbn[0]).replace('x', ''))]
            if len(str(isbn[0])) == 13:
                isbn_9 = str(isbn[0])[3:-1]
            elif len(str(isbn[0])) == 12:
                isbn_9 = str(isbn[0])[3:]
            elif len(str(isbn[0])) == 10:
                isbn_9 = str(isbn[0])[:-1]
            elif len(str(isbn[0])) == 9:
                isbn_9 = str(isbn[0])
            else:
                isbn_9 = None

            if main_category[0] != None:
                #  make the main category list in main_categories
                main_category = main_category[0].rstrip().lstrip()
                if  main_category not in main_categories:
                    main_categories.append(main_category)

                if check == 'add':
                    # grab the category object
                    category_object = cache.get('category_object_{}'.format(main_category))
                    if not category_object:
                        try:
                            category_object = Category.objects.get(name=main_category, is_main=True)
                        except:
                            category_object = Category.objects.create(
                                name = main_category,
                                is_main = True
                            )
                        cache.set('category_object_{}'.format(main_category), category_object)


                #  grab the database object
                try:
                    # try with isbn_9
                    database_product = Product.objects.filter(available=True).exclude(product_type='craft').get(isbn_9=isbn_9)
                    if check == 'add':
                        database_product.category = category_object
                        database_product.save()
                except:
                    pass

                # if not database_product:
                #     #  try with isbn
                #     try:
                #         database_product = Product.objects.filter(available=True).get(isbn=str(isbn[0]))
                #         if check == 'add':
                #             database_product.category = category_object
                #             database_product.save()
                #     except:
                #         pass

                if not database_product:
                    try:
                        #  try with name
                        database_product = Product.objects.filter(available=True).exclude(product_type='craft').get(name=name[0])
                    except MultipleObjectsReturned:
                        duplicate_name.append((name[0], publisher[0], pages[0]))

                    except ObjectDoesNotExist:
                        pass


                if check == 'add':
                    if database_product:
                        database_product.category = category_object
                        database_product.save()
                    if database_products:
                        database_products = Product.objects.filter(available=True).exclude(product_type='craft').filter(name=name[0], publisher=publisher[0])
                        database_products.update(category=main_category)

                #  maybe we could not find the book
                if not database_product:
                    not_found_product.append((name[0], isbn[0]))


            if sub_category[0] != None:
                sub_category = sub_category[0].rstrip().lstrip()
                if len(sub_category) == 1:
                    sub_category = main_category
                if sub_category not in sub_categories:
                    sub_categories.append(sub_category)

                if database_product and check=='add':
                    sub_category_object = cache.get('sub_category_object_{}'.format(sub_category))
                    if not sub_category_object:
                        try:
                            sub_category_object = Category.objects.get(name=sub_category, parent_category__id=category_object.id)
                        except:
                            sub_category_object = Category.objects.create(
                                name = sub_category,
                                parent_category = category_object,
                                is_sub = True
                            )
                        cache.set('sub_category_object_{}'.format(sub_category), sub_category_object)

                    database_product.sub_category = sub_category_object
                    database_product.save()






    main_categories = sorted(main_categories)
    sub_categories = sorted(sub_categories)
    print('main_categories', len(main_categories))
    print('sub_categories', len(sub_categories))
    # print('no_isbn', len(no_isbn))
    print('not_found_product', len(not_found_product))

    return render(
        request,
        'files/check_category.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'main_categories': main_categories,
            'sub_categories': sub_categories,
            'not_found_product': not_found_product,
            'duplicate_name': duplicate_name,

         }
    )



def update_publishers(request, file_slug, check='check'):
    final = []


    # excel file handle
    file_object = get_object_or_404(FileObject, slug=file_slug)

    myfile = File(file_object)

    path = file_object.file.path
    with open(path, 'rb') as f:
        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 45):
            publisher_2 = None

            # row = ws['A'+str(i):'M'+str(i)]

            false_name = ws['B' + str(i)].value,
            pub_1 = ws['C' + str(i)].value,
            pub_2 = ws['D' + str(i)].value,

            # print(false_name, pub_1, pub_2)
            if str(false_name[0]) in ('-', '0', 'None'):
                continue

            databse_publisher_1 = cache.get('publisher_{}'.format(str(pub_1[0])))
            if not databse_publisher_1:
                try:
                    databse_publisher_1 = Publisher.objects.get(name=pub_1[0])
                except:
                    databse_publisher_1 = Publisher.objects.create(
                        name = pub_1[0],
                    )
                cache.set('publisher_{}'.format(str(pub_1[0])), databse_publisher_1)

            # print(pub_2[0], type(pub_2[0]))
            if pub_2[0] not in ('', None):
                publisher_2 = cache.get('publisher_{}'.format(str(pub_2[0])))
                if not publisher_2:
                    try:
                        publisher_2 = Publisher.objects.get(name=pub_2[0])
                    except:
                        publisher_2 = Publisher.objects.create(
                            name = pub_2[0],
                        )
                    cache.set('publisher_{}'.format(str(pub_2[0])), publisher_2)

            products_counts = Product.objects.filter(publisher=false_name[0]).count()
            final.append((false_name[0], products_counts, databse_publisher_1, publisher_2))

            if check == 'add':
                products = Product.objects.filter(publisher=false_name[0])
                if publisher_2:
                    products.update(pub_1=databse_publisher_1, pub_2=publisher_2)
                else:
                    products.update(pub_1=databse_publisher_1)


    return render(
        request,
        'files/update_publishers.html',
        {
            'file': myfile,
            'file_object': file_object,
            'row_count': row_count,
            'final': final,
         }
    )
