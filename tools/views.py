from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from django.conf import settings
from django.http import HttpResponse, FileResponse
from django.template.loader import render_to_string
from django.contrib.admin.views.decorators import staff_member_required
from io import BytesIO
from django.db.models import Q
from datetime import datetime, timedelta
from django.core.mail import EmailMessage, mail_admins, mail_managers, get_connection
import random
from django.core.files.base import ContentFile
from django.core.files import File
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db.models import Sum, Count


import weasyprint
import openpyxl
from openpyxl.styles import Color, Fill
import qrcode
import qrcode.image.svg
from tools.gregory_to_hijry import hij_strf_date, greg_to_hij_date

from orders.models import Order, OrderLine
from zarinpal.models import Payment
from tools.gregory_to_hijry import hij_strf_date, greg_to_hij_date
from shop.models import Slogan
from account.models import CustomUser
from products.models import Product

@staff_member_required
def make_invoice_pdf(request, order_id=None):
    order = Order.objects.get(pk=order_id)

    fa_date = hij_strf_date(greg_to_hij_date(order.created.date()), '%-d %B %Y')

    html = render_to_string('tools/pdf/invoice_pdf.html',
        {'order': order,
        'fa_date': fa_date})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename=order_{order.id}.pdf'
    weasyprint.HTML(string=html).write_pdf(response,
        stylesheets=[weasyprint.CSS(settings.STATIC_ROOT + 'css/invoice_pdf.css')])
    return response


@staff_member_required
def print_invoice(request, order_id=None):
    order = Order.objects.get(pk=order_id)
    slogan_id = random.choice(Slogan.objects.values_list('pk', flat=True))
    slogan = Slogan.objects.get(pk=slogan_id)
    fa_date = hij_strf_date(greg_to_hij_date(order.created.date()), '%-d %B %Y')

    return render(
        request,
        'tools/pdf/print_invoice.html',
        {
            'order': order,
            'fa_date': fa_date,
            'slogan': slogan,
        }
    )


@staff_member_required
def print_address(request, client_id, kind='billing'):
    client = CustomUser.objects.get(pk=client_id)

    return render(
        request,
        'tools/pdf/print_address.html',
        {
            'client': client,
            'kind': kind
        }
    )


@staff_member_required
def make_invoice_pdf_a4(request, order_id=None):
    order = Order.objects.get(pk=order_id)

    fa_date = hij_strf_date(greg_to_hij_date(order.created.date()), '%-d %B %Y')

    html = render_to_string('tools/pdf/invoice_pdf_a4.html',
        {'order': order,
        'fa_date': fa_date})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename=order_{order.id}.pdf'
    weasyprint.HTML(string=html).write_pdf(response,
        stylesheets=[weasyprint.CSS(settings.STATIC_ROOT + 'css/invoice_pdf_a4.css')])
    return response


def order_export_excel(request, criteria, date=None):
    # TODO: we handle all order export with one function but we should add parameter to
    # handle differnt kind of report
    orders = None
    filename = None
    if date:
        date = datetime.strptime(date, "%Y%m%d").date()
        try:
            date = datetime.strptime(date, "%Y%m%d").date()
        except ValueError:
            messages.error(request, _('Date is not valid'))
            return redirect('staff:full_shipped_list')
        fa_date = hij_strf_date(greg_to_hij_date(date), '%-d %B %Y')
        orders = Order.objects.filter(active=True).filter(full_shipped_date__date=date).filter(full_shipped_date__isnull=False).filter(shipping_method='bike_delivery')
    else:
        if criteria in ('draft', 'approved' ):
            # we use orders list in session to export only nedded orders
            orders = Order.objects.filter(id__in=request.session['orders'])
            print(orders.count())
        elif criteria == 'full':
            orders = Order.objects.filter(active=True).filter(full_shipped_date__isnull=False).filter(shipping_method='bike_delivery')

    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        'Order ID',
        'Client',
        'Client Phone',
        'Created',
        'Created by',
        'Approved date',
        'Approved by',
        'Status',
        'Channel',
        'Billing address',
        'Shipping address',
        'Shipping method',
        'Shipping cost',
        'Shipping time note',
        'Shipping status',
        'Shipped code',
        'Full shipped date',
        'Is packaged',
        'Total cost',
        'Total cost after discount',
        'Order discount',
        'Payable',
        'Is paid',
        'Paid by credit',
        'Customer notes',
        'Quantity',
        'Weight',
        'Is gift',
    ]

    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]

    # writing body
    for count , order in enumerate(orders):
        order.save()
        title_list = [
            order.id,
            str(order.client),
            str(order.client_phone),
            str(order.created,),
            str(order.user),
            str(order.approved_date),
            str(order.approver),
            order.status,
            order.channel,
            order.billing_address.get_full_address() if order.billing_address else '',
            order.shipping_address.get_full_address() if order.shipping_address else '',
            order.shipping_method,
            order.shipping_cost,
            order.shipping_time,
            order.shipping_status,
            order.shipped_code if order.shipped_code else '',
            hij_strf_date(greg_to_hij_date(order.full_shipped_date.date()), '%-d %B %Y') if order.full_shipped_date else '',
            order.is_packaged,
            order.total_cost,
            order.total_cost_after_discount,
            order.discount,
            order.payable,
            order.paid,
            order.pay_by_credit,
            order.customer_note,
            order.quantity,
            order.weight,
            order.is_gift,
        ]
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]

    if criteria in ('draft', 'approved' ):
        filename = 'media/excel/approved-orders-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    elif criteria == 'full':
        filename = 'media/excel/full-shipped-orders-{}.xlsx'.format(datetime.now().isoformat(sep='-'))

    wb.save(filename)

    request.session['orders'] = None

    excel = open(filename, 'rb')
    response = FileResponse(excel)
    return response


def draft_order_export_excel(request):
    # TODO: Should mix with order_export_excel
    orders = Order.objects.filter(status='draft').filter(active=True)

    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        'Order ID', 'Client', 'Client Phone', 'Created', 'Status', 'Channel',
        'billing_address', 'shipping_address', 'shipping_method',  'Created by',
        'Total cost', 'Tota cost after discount', 'discount', 'Payable', 'paid',
        'Customer notes', 'Weight', 'Is gift',
    ]

    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]


    for count , order in enumerate(orders):
        title_list = [ order.id, str(order.client), str(order.client_phone),
             str(order.created,), order.status, order.billing_address, order.shipping_address,
            order.shipping_method, str(order.user), order.total_cost, order.total_cost_after_discount, order.discount,
            order.payable, order.paid, order.customer_note, order.weight, order.is_gift,
        ]
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]


    filename = 'media/excel/draft-orders-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    wb.save(filename)
    excel = open(filename, 'rb')
    response = FileResponse(excel)


    return response


def email_to_admin(paymaent_id):
    """
    Function to send email to admin when a payment is done.
    """
    payment = Payment.objects.get(pk=paymaent_id)
    order_number = '-'
    if payment.order:
        order_number = payment.order.pk

    #email body
    subject = '[Damavand] Successful payment: {} order:{}'.format(payment.id, order_number)
    # message = 'A successful payment registered at Zarinpal: \nPayment ID.: {} \nName: {} \nPhone: {} \nAmount: {:,} \n Date & Time:{} \n Ref ID.:{}'.format(
    #     payment.id,
    #     payment.client_name,
    #     payment.client_phone,
    #     payment.amount,
    #     payment.created,
    #     payment.ref_id)
    message = f"Successful payment at Zarinpal: \nPayment ID.: {payment.id} \nOrder ID.: {order_number} \nAmount: {payment.amount:,} \n\nName: {payment.client_name} \nPhone: {payment.client_phone} \n\nDate & Time:{payment.created.isoformat(sep='-')} \n Ref ID.:{payment.ref_id}"

    # send email
    # mail_managers(
    #     subject,
    #     message,
    #     fail_silently=False
    #     )
    connection = get_connection()
    connection.open()
    admin_email = EmailMessage(
        subject,
        message,
        'noreply@ketabedamavand.com',
        [
        'rahim.aghareb@gmail.com',
        'bahman.shafaa@gmail.com'
        ]
    )
    admin_email.send()
    connection.close()


def notif_email_to_managers(subject, message, recivers):
    """
    Function to send notification email to admin when it's needed.
    """

    #email body
    # subject = string
    # message = string
    # recivers = list()

    connection = get_connection()
    connection.open()
    admin_email = EmailMessage(
        subject,
        message,
        'noreply@ketabedamavand.com',
        recivers
    )
    admin_email.send()
    connection.close()


def product_export_excel(request, filter='all'):

    # check the ordelines for sold product
    orderline_id_list = OrderLine.objects.filter(active=True).exclude(product__product_type='craft').values_list('product__id', flat=True)

    # if statement to export two kind of excel file with one code block
    if filter == 'used-noprice':
        products = Product.objects.filter(available=True).exclude(product_type='craft').filter(stock_used__gte=1).filter(price_used=0).order_by('name')
    elif filter == 'used-all':
        products = Product.objects.filter(available=True).exclude(product_type='craft').filter(Q(stock_used__gte=1) | Q(price_used__gte=1)).order_by('name')
    else:
        # export all product books and crafts
        products = Product.objects.filter(available=True).exclude(product_type='craft').order_by('name')

    wb = openpyxl.Workbook()
    sheet = wb.active
    if filter in ('used-noprice', 'used-all'):
        headers = [
            '#',
            'Product ID.',
            'Name',
            'isbn',
            'Publisher',
            'Main price',
            'Main stock',
            'Price used',
            'Stock used',
            'Not in market',
            'Page number',
            'Weight',
            'Sold in order No.',
            'Sold quantity',
        ]
    else:
        headers = [
            '#',
            'Product ID.',
            'Name',
            'isbn',
            'Publisher',
            'Stock',
            'Price',
            'Weight',
            'size',
            'Cover type',
            'Page number',
            'Edition',
            'Other prices',
            'Price 1',
            'Stock 1',
            'Price 2',
            'Stock 2',
            'Price 3',
            'Stock 3',
            'Price 4',
            'Stock 4',
            'Price 5',
            'Stock 5',
            'Price used',
            'Stock used',
            'Not in market',
            'Available online',
            'Available in store',
            'Category',
            'Author',
            'Translator',
            'Publisher 2',
            'Publish year',
            'Product Latin name',
            'Author latin name',
            'Age range',
            'Is collection',
            'Admin note',
            'Description',
            'Import session',
        ]

    # writing header
    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]
        # c.style.fill.fill_type = Fill
        # c.style.fill.start_color.index = Color.BLUE


    # writing body
    for count , product in enumerate(products.iterator()):
        # product.save()
        if filter in ('used-noprice', 'used-all'):
            # check wether the product is sold or not
            sold_order_list = ''
            sold_quantity = 0
            if product.id in orderline_id_list:
                sold_order_list = OrderLine.objects.filter(product__id=product.id).values_list('order__id', flat=True)
                sold_order_list = ','.join([str(item) for item in sold_order_list])
                sold_quantity = OrderLine.objects.filter(product__id=product.id).aggregate(total=Sum('quantity'))['total']

            title_list = [
                count,
                product.id,
                str(product),
                product.isbn,
                str(product.pub_1),
                product.price,
                product.stock,
                product.price_used,
                product.stock_used,
                True if product.about=='*' else False,
                product.page_number,
                product.weight,
                sold_order_list,
                sold_quantity,
            ]
        else:
            title_list = [
                count,
                product.id,
                str(product),
                product.isbn,
                str(product.pub_1),
                product.stock,
                product.price,
                product.weight,
                product.size,
                product.cover_type,
                product.page_number,
                product.edition,
                product.has_other_prices,
                product.price_1,
                product.stock_1,
                product.price_2,
                product.stock_2,
                product.price_3,
                product.stock_3,
                product.price_4,
                product.stock_4,
                product.price_5,
                product.stock_5,
                product.price_used,
                product.stock_used,
                True if product.about=='*' else False,
                product.available_online,
                product.available_in_store,
                str(product.category),
                product.author,
                product.translator,
                str(product.pub_1),
                product.publish_year,
                product.latin_name,
                product.author_latin_name,
                product.age_range,
                product.is_collection,
                product.admin_note,
                product.description,
                str(product.import_session),
            ]
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]

    if filter=='used-noprice':
        filename = 'media/excel/used-noprice-products-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    elif filter == 'used-all':
        filename = 'media/excel/used-all-products-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    else:
        filename = 'media/excel/available-products-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    wb.save(filename)
    excel = open(filename, 'rb')
    response = FileResponse(excel)


    return response


def used_product_before_5(request):
    point = datetime.strptime('2022 3 5 16 11 26', "%Y %m %d %H %M %S")
    products = OrderLine.objects.filter(active=True).filter(created__lte=point).filter(variation__contains='used')

    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        '#',
        'Product ID.',
        'Name',
        'isbn',
        'Publisher',
        'Main price',
        'Main stock',
        'Price used',
        'Stock used',
        'OrderLine quantity',
        'date',
        'History',
        'Sold quantity',
    ]
    # writing header
    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]
        # c.style.fill.fill_type = Fill
        # c.style.fill.start_color.index = Color.BLUE


    # making body
    for count , item in enumerate(products.iterator()):
        item_history = item.product.history.all()
        # item_history_as = item.product.history.as_of(point)
        sold_quantity = OrderLine.objects.filter(product__id=item.product.id).filter(variation__contains='used').aggregate(total=Sum('quantity'))['total']
        # print(type(item_history_as), item_history_as)
        history_list =[]
        for line in item_history:

            history = f"{line.stock_used}"
            history_list.append(history)

        title_list = [
            count,
            item.product.id,
            str(item.product),
            item.product.isbn,
            item.product.pub_1.name,
            item.product.price,
            item.product.stock,
            item.product.price_used,
            item.product.stock_used,
            item.quantity,
            item.created.date().strftime("%Y%m%d"),
            ','.join(history_list),
            sold_quantity
        ]
        # print (item.product.id, item.product.name)
        # writing body
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]

    filename = 'media/excel/used-before-5-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    wb.save(filename)
    excel = open(filename, 'rb')
    response = FileResponse(excel)
    return response

    # return redirect('staff:products')
    # return render (
    #     request,
    #     'staff/_used_before_5.html',
    #     {
    #         'products': products,
    #     }
    # )


def qrcode_create(request, order_id, payment_id):
    order = get_object_or_404(Order, pk=order_id)
    payment = get_object_or_404(Payment, pk=payment_id)
    context = {}
    factory = qrcode.image.svg.SvgImage
    img = qrcode.make(payment.url, image_factory=factory, box_size=20)
    stream = BytesIO()
    # img.save(stream, img.format)
    img.save(stream)

    # image.save(strem)
    # context['svg'] = stream.getvalue().decode()
    # return render('staff:order_checkout', order.id)
    f = open('media/test.svg', 'w')
    f.write(stream.getvalue().decode())
    f.close()
    order.qrcode = f
    order.save()

    return render(
        request,
        'tools/qrcode_test.html',
        {
            'svg': stream.getvalue().decode()
            # 'svg': img
        }
    )


@staff_member_required
def export_publisher(request):
    # TODO: we handle all order export with one function but we should add parameter to
    # handle differnt kind of report
    products = Product.objects.filter(available=True).exclude(product_type='craft').exclude(publisher='').order_by('publisher').distinct('publisher')

    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        '#',
        'Publisher',
    ]

    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]


    for count , product in enumerate(products):

        title_list = [
            count,
            product.pub_1.name,
        ]
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]


    filename = 'media/excel/publishers-{}.xlsx'.format(datetime.now().isoformat(sep='-'))
    wb.save(filename)
    excel = open(filename, 'rb')
    response = FileResponse(excel)

    return response


@staff_member_required
def export_excel_sold_products(request, date=None, days=None, period=None):
    fa_date = None
    if days:
        order_lines = OrderLine.objects.all().filter(active=True).filter(created__gte=datetime.now() - timedelta(days)).exclude(product__product_type='craft').order_by('-created')
    if date:
        date = date.replace('-', '')
        date = datetime.strptime(date, "%Y%m%d").date()
        fa_date = hij_strf_date(greg_to_hij_date(date), '%-d %B %Y')
        order_lines = OrderLine.objects.all().filter(active=True).filter(created__date=date).exclude(product__product_type='craft').order_by('-created')

    if period:
        period_list = period.split('-')
        start = datetime.strptime(period_list[0], "%Y%m%d").date()
        end = datetime.strptime(period_list[1], "%Y%m%d").date()
        order_lines = OrderLine.objects.all().filter(active=True).filter(
            created__date__range=(start, end)).exclude(product__product_type='craft').order_by('-created')
    products = dict()
    main_stock = dict()
    check_list =[]
    added_line = {}
    # added_line = {
    #     'pk': {
    #         'quantity': 'x',
    #         'name': 'str',
    #         'created': 'created',
    #         'stock': 'y',
    #          'variation': 'new main'
    #          'id': 'id'
    #     }
    # }

    for item in order_lines:
        # key = products.get(item.product.id)
        # print(item.product.id)
        # if item.product.name in added_line:
        if str(item.product.id) in added_line:
            added_line[str(item.product.id)]['quantity'] += item.quantity
            pass
        else:
            # messages.warning(request, 'hi')
            # check_list.append(item.product.name)
            added_line[str(item.product.id)] = {
                'name': str(item.product),
                'quantity': item.quantity,
                'created': hij_strf_date(greg_to_hij_date(item.created.date()), '%-d %B %Y'),
                'stock': item.product.stock,
                'variation': item.variation,
                'isbn': item.product.isbn,
                'id': item.product.id,
                'publisher': item.product.pub_1.name,
                'vendors': ','.join(item.product.vendors.all().values_list('first_name', flat=True)),
                }

    # print(added_line)
    # Create excel file
    wb = openpyxl.Workbook()
    sheet = wb.active

    # create headers
    headers = [
        fa_date if fa_date else days,
        'Product ID',
        'ISBN',
        'Product',
        'Sold quantity',
        'Variation',
        'Main stock',
        'Publisher',
        'Vendors',
    ]
    # write headers
    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]

    body_list = list(added_line.keys())
    # print(body_list)
    # create body
    for count , key in enumerate(body_list):
        title_list = [
            count,
            added_line[key]['id'],
            added_line[key]['isbn'],
            added_line[key]['name'],
            added_line[key]['quantity'],
            'Used' if 'used' in added_line[key]['variation'] else 'New',
            added_line[key]['stock'],
            added_line[key]['publisher'],
            added_line[key]['vendors'],
        ]
        # write body
        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]

    if date:
        filename = 'media/excel/sold-products-by-{}-{}.xlsx'.format(date, datetime.now().isoformat(sep='-'))
    if days:
        filename = 'media/excel/sold-products-by-{}-{}.xlsx'.format(days, datetime.now().isoformat(sep='-'))
    if period:
        filename = 'media/excel/sold-products-by-{}-{}.xlsx'.format(period, datetime.now().isoformat(sep='-'))

    wb.save(filename)
    excel = open(filename, 'rb')
    response = FileResponse(excel)

    return response


@staff_member_required
def duplicate_book_name(request):

    from openpyxl import load_workbook
    from files.models import File as FileObject

    file_object = get_object_or_404(FileObject, slug='leven-list')
    path = file_object.file.path

    similar_ids = []

    with open(path, 'rb') as f:

        # Load excel workbook
        wb = load_workbook(f)
        ws = wb.active
        row_count = ws.max_row

        # a loop for scrape the excel file
        # read the data in cells
        for i in range(2, row_count):
        # for i in range(2, 15):

            row = ws['A'+str(i):'E'+str(i)]

            product_id_1 = ws['C' + str(i)].value,
            product_id_2 = ws['E' + str(i)].value,

            similar_ids.append(product_id_1[0])
            similar_ids.append(product_id_2[0])
    # similars are those we find with lebenshtein distance
    similar_products = Product.objects.filter(pk__in=similar_ids)
    similar_products_count = Product.objects.filter(pk__in=similar_ids).count()

    # these are the products with the exact same name
    products_list = Product.objects.filter(available=True).exclude(product_type='craft').values('name').annotate(Count('id')).order_by().filter(id__count__gt=1)
    products_list_count = len(products_list)

    same_products = Product.objects.filter(name__in=[item['name'] for item in products_list]).order_by('name')
    same_products_count = same_products.count()

    products = same_products | similar_products
    products_count = products.count()

    products = products.order_by('name')

    request.session['duplicate_book_name'] = list(products.values_list('id', flat=True))

    return render(
        request,
        'staff/duplicate_book_name.html',
        {
            'products': products,
            'products_list_count': products_list_count,
            'products_count': products_count,
            'similar_products_count': similar_products_count,
            'same_products_count': same_products_count,
        }
    )


@staff_member_required
def duplicate_book_name_export(request):

    products = Product.objects.filter(id__in=request.session['duplicate_book_name']).order_by('name')


    # check the ordelines for sold product
    orderline_id_list = OrderLine.objects.filter(active=True).exclude(product__product_type='craft').values_list('product__id', flat=True)

    # if statement to export two kind of excel file with one code block
    # if filter == 'used-noprice':
    #     products = Product.objects.filter(available=True).exclude(product_type='craft').filter(stock_used__gte=1).filter(price_used=0).order_by('name')
    # elif filter == 'used-all':
    #     products = Product.objects.filter(available=True).exclude(product_type='craft').filter(Q(stock_used__gte=1) | Q(price_used__gte=1)).order_by('name')
    # else:
    #     # export all product books and crafts
    #     products = Product.objects.filter(available=True).exclude(product_type='craft').order_by('name')

    wb = openpyxl.Workbook()
    sheet = wb.active

    headers = [
        '#',
        'Product ID.',
        'Name',
        'isbn',
        'Publisher',
        'Main price',
        'Main stock',
        'Price used',
        'Stock used',
        'Not in market',
        'Page number',
        'Weight',
        'Sold in order No.',
        'Sold quantity',
    ]


    # writing header
    for i in range(len(headers)):
        c = sheet.cell(row = 1, column = i + 1 )
        c.value = headers[i]
        # c.style.fill.fill_type = Fill
        # c.style.fill.start_color.index = Color.BLUE


    # writing body
    for count , product in enumerate(products.iterator()):
        # product.save()
        # if filter in ('used-noprice', 'used-all'):
        # check wether the product is sold or not
        sold_order_list = ''
        sold_quantity = 0
        if product.id in orderline_id_list:
            sold_order_list = OrderLine.objects.filter(product__id=product.id).values_list('order__id', flat=True)
            sold_order_list = ','.join([str(item) for item in sold_order_list])
            sold_quantity = OrderLine.objects.filter(product__id=product.id).aggregate(total=Sum('quantity'))['total']

        title_list = [
            count,
            product.id,
            str(product),
            product.isbn,
            str(product.pub_1),
            product.price,
            product.stock,
            product.price_used,
            product.stock_used,
            True if product.about=='*' and product.price==0 else False,
            product.page_number,
            product.weight,
            sold_order_list,
            sold_quantity,
        ]

        for i in range(len(headers)):
            c = sheet.cell(row = count + 2 , column = i + 1)
            c.value = title_list[i]


    filename = 'media/excel/duplicate-products-{}.xlsx'.format(datetime.now().isoformat(sep='-'))

    wb.save(filename)

    request.session['duplicate_book_name'] = None

    excel = open(filename, 'rb')
    response = FileResponse(excel)


    return response
